from tkinter import *
from datetime import date, datetime
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib
import tkinter as Tk
# import subprocess
import random,os,tempfile,smtplib
from tkinter import Label, Entry, ttk, Scrollbar, Frame, VERTICAL, Canvas
from email.mime.text import MIMEText


# --- Funtionality part --- 


# --- tao folder bill ----
if not os.path.exists('bills'):
    os.mkdir('bills')

# --- ham save Bill -----
def save_bill():
    global billnumber
    result=messagebox.askyesno('Confirm''Do you want to save the bill ?')
    if result:
        bill_content=textarea.get(1.0,END)
        file=open(f'bills/ {billnumber}.txt','w')
        file.write(bill_content)
        file.close()
        messagebox.showinfo('Success'f'bill number {billnumber} is saved successfully')
        billnumber=random.randint(500,1000)

billnumber=random.randint(500,1000)

    
# ----- Gọi cửa sổ tính tiền ----
   
    
def open_recapp():    
    root0 = Toplevel(root)
    root0.grab_set()   
    # root.grab_set()
    root0.title('Rehnung System')
    root0.geometry('1250x730')
    root0.iconbitmap('Images/bill.ico')
    background = "#06283D"  
    root0.config(bg=background)
   
    # --- Exit ----
    def Exitbutton():
        root0.destroy() 
    # --- Clear -----------
    def clear():
        RoseDayEntry.delete(0,END)
        RoseNightEntry.delete(0,END)
        WhiteRoseDayEntry.delete(0,END)
        WhiteRoseNightEntry.delete(0,END)
        BodylotionEntry.delete(0,END)
        HandCreamEntry.delete(0,END)
        RoseWaterEntry.delete(0,END)
        
        HaarefarbeEntry.delete(0,END)
        DamenhaareEntry.delete(0,END)
        DamenföhlenEntry.delete(0,END)
        Cu_wa_foEntry.delete(0,END)
        WimpernVEntry.delete(0,END)
        HairCutMEntry.delete(0,END)
        HairWashMEntry.delete(0,END)
        
        AuffüllenEntry.delete(0,END)
        NeueSetEntry.delete(0,END)
        ShellachEntry.delete(0,END)
        BedechungEntry.delete(0,END)
        PedikureEntry.delete(0,END)
        ManikureEntry.delete(0,END)
        AcrylEntEntry.delete(0,END)
        
        # -- tra lai cac o san pham 0 ---
        RoseDayEntry.insert(0,0)
        RoseNightEntry.insert(0,0)
        WhiteRoseDayEntry.insert(0,0)
        WhiteRoseNightEntry.insert(0,0)
        BodylotionEntry.insert(0,0)
        HandCreamEntry.insert(0,0)
        RoseWaterEntry.insert(0,0)
        
        HaarefarbeEntry.insert(0,0)
        DamenhaareEntry.insert(0,0)
        DamenföhlenEntry.insert(0,0)
        Cu_wa_foEntry.insert(0,0)
        WimpernVEntry.insert(0,0)
        HairCutMEntry.insert(0,0)
        HairWashMEntry.insert(0,0)
        
        AuffüllenEntry.insert(0,0)
        NeueSetEntry.insert(0,0)
        ShellachEntry.insert(0,0)
        BedechungEntry.insert(0,0)
        PedikureEntry.insert(0,0)
        ManikureEntry.insert(0,0)
        AcrylEntEntry.insert(0,0)
        
        NameEntry.delete(0,END)
        PhoneEntry.delete(0,END)
        billnumberEntry.delete(0,END)
        ProductspriceEntry.delete(0,END)
        HaarepriceEntry.delete(0,END)
        NagelpriceEntry.delete(0,END)
        ProductstaxEntry.delete(0,END)
        HaaretaxEntry.delete(0,END)
        NageltaxEntry.delete(0,END)
        textarea.delete(1.0,END)
        
    # --------- Gửi Email - chưa được sử lý ----
    def send_email():
        
            
        def send_billMail():
            try:
                ob=smtplib.SMTP('smtp.gmail.com',587)
                ob.starttls()
                ob.login(senderEntry.get(),passwordEntry.get())
                message=email_textarea.get(1.0,END)
                ob.sendmail(senderEntry.get(),recieverEntry.get(),message)
                ob.quit()
                messagebox.showinfo('Success','Bill is successfully send',parent=root1)
                root1.destroy()
            except smtplib.SMTPAuthenticationError:
                messagebox.showerror('Error', 'Authentication Error! Check your email and password.', parent=root1)
            except smtplib.SMTPException as e:
                messagebox.showerror('Error', f'Something went wrong: {str(e)}', parent=root1)
        if textarea.get(1.0, END) == '\n':
            messagebox.showerror('Error', 'Bill is empty!', parent=root0)       
        else:   
            root1=Toplevel(root0)
            root1.grab_set()
            root1.title('Send Email')
            root1.config(bg='#06283D')
            root1.resizable(0,0)
            root1.iconbitmap('Images/bill.ico')
            
            senderFrame=LabelFrame(root1, text='Sender', font=('Times new roman',16,'bold'),fg='#06283D',bd=6)
            senderFrame.grid(row=0,column=0,padx=20,pady=10)
                
            senderLabel=Label(senderFrame, text="Sender's Email: ", font=('Times new roman',12,'bold'),fg='#06283D')
            senderLabel.grid(row=0,column=0,padx=10,pady=8)
            senderEntry=Entry(senderFrame, font=('Times new roman',12,'bold'),bd=2,width=28,relief=RIDGE)
            senderEntry.grid(row=0,column=1,padx=10,pady=8)
                
            passwordLabel=Label(senderFrame, text="Sender's Password: ", font=('Times new roman',12,'bold'),fg='#06283D')
            passwordLabel.grid(row=1,column=0,padx=10,pady=8)
            passwordEntry=Entry(senderFrame, font=('Times new roman',12,'bold'),bd=2,width=28,relief=RIDGE)# ,show='*'
            passwordEntry.grid(row=1,column=1,padx=10,pady=8)
                
            recipientFrame=LabelFrame(root1, text='recipient', font=('Times new roman',16,'bold'),fg='#06283D',bd=6)
            recipientFrame.grid(row=1,column=0,padx=20,pady=10)
                
            recieverLabel=Label(recipientFrame, text="reciever's Email: ", font=('Times new roman',12,'bold'),fg='#06283D')
            recieverLabel.grid(row=0,column=0,padx=10,pady=8)
            recieverEntry=Entry(recipientFrame, font=('Times new roman',12,'bold'),bd=2,width=28,relief=RIDGE)
            recieverEntry.grid(row=0,column=1,padx=10,pady=8)
                
            messengeLabel=Label(recipientFrame, text="messenge:", font=('Times new roman',12,'bold'),fg='#06283D')
            messengeLabel.grid(row=1,column=0,padx=10,pady=8)
            email_textarea=Text(recipientFrame, font=('Times new roman',12,'bold'),bd=2, relief=SUNKEN, width=50,height=20)
            email_textarea.grid(row=2,column=0,columnspan=2)
            email_textarea.delete(1.0,END)
            email_textarea.insert(END,textarea.get(1.0,END).replace('=','').replace('-',''))
                
            sendButton=Button(root1, text='SEND', font=('Times new roman',12,'bold'),width=15,bd=3,command=send_billMail)
            sendButton.grid(row=2,column=0,pady=20)
            
            
            root1.mainloop()
    
    # -- in hóa đơn --code này chạy không lỗi và in ra giấy --
    def print_bill():
        if textarea.get(1.0, END) == '\n':
            messagebox.showerror('Error', 'Bill is empty!')
        else:
            temp_dir = tempfile.gettempdir()  # Get the system's default temp directory
            file_path = os.path.join(temp_dir, 'bill.txt')  # Create a unique path with filename
            with open(file_path, 'w') as f:
                f.write(textarea.get(1.0, END))
            os.startfile(file_path, 'print')
    
    #--- ham tim kiem ----
    def search_bills():
        for i in os.listdir('bills/'):
            
            if i.split('.')[0].strip() == billnumberEntry.get().strip():
                f = open(f'bills/{i}', 'r')
                textarea.delete(1.0, END)
                for data in f: 
                    textarea.insert(END, data) 
                f.close()
                break
        else:
            messagebox.showerror('Error','Invalid bill Number !', parent=root0)

# --- tao folder bill ----
    if not os.path.exists('bills'):
        os.mkdir('bills')

# --- ham save Bill -----
    def save_bill():
        pass
        global billnumber
        result=messagebox.askyesno('Confirm''Do you want to save the bill ?', parent=root0)
        if result:
            bill_content=textarea.get(1.0,END)
            file=open(f'bills/ {billnumber}.txt','w')
            file.write(bill_content)
            file.close()
            messagebox.showinfo('Success'f'bill number {billnumber} is saved successfully', parent=root0)
            billnumber=random.randint(500,1000)

    billnumber=random.randint(500,1000)

# ---- Bill Area ------
    def bill_Area():
        if NameEntry.get()=='' or PhoneEntry.get()=='':
            messagebox.showerror('Error','Customer Name and Phon Number Required !')
        elif ProductspriceEntry.get()=='' or HaarepriceEntry.get()=='' or NagelpriceEntry.get()=='':
            messagebox.showerror('Error','Products are empty !')
        elif ProductspriceEntry.get()=='0 Euro' or HaarepriceEntry.get()=='0 Euro' or NagelpriceEntry.get()=='0 Euro':
            messagebox.showerror('Error','Products are empty !')
            
            textarea.delete(1.0,END)
        else:
            textarea.insert(END,'\t***Wellcome Customer***\nCosmetic Shop\nWiener Strasse 61\nA-4020 Linz\n')
            textarea.insert(END,'----------------\n')
            textarea.insert(END,f'Customer Name: {NameEntry.get()}\n')
            textarea.insert(END,f'Phone Number: {PhoneEntry.get()}\n')
            textarea.insert(END,f'Bill Number: {billnumber}\t')
            textarea.insert(END,f'\tDate: {DateEntry.get()}\n')
            textarea.insert(END,'=================================================\n')
            textarea.insert(END,'Products \t\tQuantity\t\tPrice\n')
            textarea.insert(END,'=================================================\n')
        if RoseDayEntry.get()!='0':
            textarea.insert(END,f'Rose Day Cream\t\t\t{RoseDayEntry.get()}\t{RoseDayprice} Euro.\n')
        if RoseNightEntry.get()!='0':
            textarea.insert(END,f'Rose Night Cream\t\t\t{RoseNightEntry.get()}\t{RoseNightprice} Euro\n')
        if WhiteRoseDayEntry.get()!='0':
            textarea.insert(END,f'White Rose Day \t\t\t{WhiteRoseDayEntry.get()}\t{WhiteRoseDayprice} Euro\n')
        if WhiteRoseNightEntry.get()!='0':
            textarea.insert(END,f'White Rose Night \t\t\t{WhiteRoseNightEntry.get()}\t{WhiteRoseNightprice} Euro\n')
        if BodylotionEntry.get()!='0':
            textarea.insert(END,f'Bodylotion\t\t\t{BodylotionEntry.get()}\t{Bodylotionprice} Euro\n')
        if HandCreamEntry.get()!='0':
            textarea.insert(END,f'Rose Hand Cream\t\t\t{HandCreamEntry.get()}\t{HandCreamprice} Euro\n')
        if RoseWaterEntry.get()!='0':
            textarea.insert(END,f'Rose Water\t\t\t{RoseWaterEntry.get()}\t{RoseWaterprice} Euro\n')
                
        if AuffüllenEntry.get()!='0':
            textarea.insert(END,f'Auffüllen\t\t\t{AuffüllenEntry.get()}\t{Auffüllenprice} Euro\n')
        if NeueSetEntry.get()!='0':
            textarea.insert(END,f'Neue Set\t\t\t{NeueSetEntry.get()}\t{NeueSetprice} Euro\n')
        if ShellachEntry.get()!='0':
            textarea.insert(END,f'Shellach\t\t\t{ShellachEntry.get()}\t{Shellachprice} Euro\n')
        if BedechungEntry.get()!='0':
            textarea.insert(END,f'Bedechung\t\t\t{BedechungEntry.get()}\t{Bedechungprice} Euro\n')
        if PedikureEntry.get()!='0':
            textarea.insert(END,f'Pedikure\t\t\t{PedikureEntry.get()}\t{Pedikureprice} Euro\n')
        if ManikureEntry.get()!='0':
            textarea.insert(END,f'Manikure\t\t\t{ManikureEntry.get()}\t{Manikureprice} Euro\n')
        if AcrylEntEntry.get()!='0':
            textarea.insert(END,f'Acryl Entfernen\t\t\t{AcrylEntEntry.get()}\t{AcrylEntprice} Euro\n')
                
        if HaarefarbeEntry.get()!='0':
            textarea.insert(END,f'Haare-farbe\t\t\t{HaarefarbeEntry.get()}\t{Haarefarbeprice} Euro\n')
        if DamenhaareEntry.get()!='0':
            textarea.insert(END,f'Damen haare cut\t\t\t{DamenhaareEntry.get()}\t{Damenhaareprice} Euro\n')
        if DamenföhlenEntry.get()!='0':
            textarea.insert(END,f'D.Waschen+föhnen\t\t\t{DamenföhlenEntry.get()}\t{Damenföhlenprice} Euro\n')
        if Cu_wa_foEntry.get()!='0':
            textarea.insert(END,f'C.W.+ Föhnen\t\t\t{Cu_wa_foEntry.get()}\t{Cu_wa_foprice} Euro\n')
        if WimpernVEntry.get()!='0':
            textarea.insert(END,f'wimpern verlängerung\t\t\t{WimpernVEntry.get()}\t{WimpernVprice} Euro\n')
        if HairCutMEntry.get()!='0':
            textarea.insert(END,f'Hair Cut Men\t\t\t{HairCutMEntry.get()}\t{HairCutMprice} Euro\n')
        if HairWashMEntry.get()!='0':
            textarea.insert(END,f'Hair Wash Men\t\t\t{HairWashMEntry.get()}\t{HairWashMprice} Euro\n')    
        textarea.insert(END,'-------------------------------------------------')
            
        if ProductstaxEntry.get()!='0.0 Euro':
            textarea.insert(END,f'\n Produkte Tax: \t\t    {ProductstaxEntry.get()}')
        if HaaretaxEntry.get()!='0.0 Euro':
            textarea.insert(END,f'\n Haarespflege Tax: \t\t{HaaretaxEntry.get()}')
        if NageltaxEntry.get()!='0.0 Euro':
            textarea.insert(END,f'\n Nagelspflege Tax: \t\t{NageltaxEntry.get()}\n')
        textarea.insert(END,f'\n\t\tTotal Bill \t\t{totalbill:.2f}\n')
        textarea.insert(END,'-------------------------------------------------\n')
        textarea.insert(END,'\t**Thank you**\t See you again')
            
        save_bill()    
           
   
    # ------------------end check out Button ---- 
    # --- Total Price Tax ----
    def Check_outbutton():
        
        global RoseDayprice,RoseNightprice,WhiteRoseDayprice,WhiteRoseNightprice,Bodylotionprice,HandCreamprice,RoseWaterprice
        global Auffüllenprice,NeueSetprice,Shellachprice,Bedechungprice,Pedikureprice,Manikureprice,AcrylEntprice
        global Haarefarbeprice, Damenhaareprice, Damenföhlenprice, Cu_wa_foprice, WimpernVprice, HairCutMprice, HairWashMprice
        global totalbill
            
        RoseDayprice=int(RoseDayEntry.get())*20
        RoseNightprice=int(RoseNightEntry.get())*20
        WhiteRoseDayprice=int(WhiteRoseDayEntry.get())*26
        WhiteRoseNightprice=int(WhiteRoseNightEntry.get())*26
        Bodylotionprice=int(BodylotionEntry.get())*10
        HandCreamprice=int(HandCreamEntry.get())*8
        RoseWaterprice=int(RoseWaterEntry.get())*16
            
        TotalProductsprice=RoseDayprice+RoseNightprice+WhiteRoseDayprice+WhiteRoseNightprice+Bodylotionprice+HandCreamprice+RoseWaterprice
        ProductspriceEntry.delete(0,END)
        ProductspriceEntry.insert(0,f'{TotalProductsprice:.2f} Euro') # (0,f'{totalProductsprice} Euro')
        ProductsTax = round(TotalProductsprice * 0.20, 2)
        ProductstaxEntry.delete(0,END)
        ProductstaxEntry.insert(0,f'{ProductsTax:.2f} Euro')# (0, str(ProductsTax)+ ' Euro')
            
        Auffüllenprice=int(AuffüllenEntry.get())*40
        NeueSetprice=int(NeueSetEntry.get())*45
        Shellachprice=int(ShellachEntry.get())*26
        Bedechungprice=int(BedechungEntry.get())*40
        Pedikureprice=int(PedikureEntry.get())*37
        Manikureprice=int(ManikureEntry.get())*20
        AcrylEntprice=int(AcrylEntEntry.get())*15
            
        TotalNagelprice = Auffüllenprice + NeueSetprice + Shellachprice + Bedechungprice + Pedikureprice + Manikureprice + AcrylEntprice
        NagelpriceEntry.delete(0,END)
        NagelpriceEntry.insert(0,f'{TotalNagelprice:.2f} Euro')
        NagelTax = round(TotalNagelprice * 0.15, 2)
        NageltaxEntry.delete(0,END)
        NageltaxEntry.insert(0,f'{NagelTax:.2f} Euro')
            
        Haarefarbeprice=int(HaarefarbeEntry.get())*50
        Damenhaareprice=int(DamenhaareEntry.get())*30
        Damenföhlenprice=int(DamenföhlenEntry.get())*15
        Cu_wa_foprice=int(Cu_wa_foEntry.get())*50
        WimpernVprice=int(WimpernVEntry.get())*80
        HairCutMprice=int(HairCutMEntry.get())*19
        HairWashMprice=int(HairWashMEntry.get())*15
            
        TotalHaareprice = Haarefarbeprice + Damenhaareprice + Damenföhlenprice + Cu_wa_foprice + WimpernVprice + HairCutMprice + HairWashMprice
        HaarepriceEntry.delete(0,END)
        HaarepriceEntry.insert(0,f'{TotalHaareprice:.2f} Euro')
        HaareTax = round(TotalHaareprice * 0.10, 2)
        HaaretaxEntry.delete(0,END)
        HaaretaxEntry.insert(0,f'{HaareTax:.2f} Euro')
            
        totalbill = TotalProductsprice + TotalHaareprice + TotalNagelprice + ProductsTax + HaareTax + NagelTax    
    
    # ############################## end of test ################################

    DetailsFrame=Frame(root0,bg="#06283D")
    DetailsFrame.pack(side=TOP)


    TopLabel=Label(DetailsFrame, text='',font=('times new roman',12), bg="#06283D", fg='#06283D')
    TopLabel.grid(row=0, padx=True)

    NameLabel=Label(DetailsFrame,text='Name',font=('times new roman',12,'bold'),bg="#06283D",fg='#fff')
    NameLabel.grid(row=1,column=0,pady=9,padx=10)
    NameEntry=Entry(DetailsFrame,font=('times new roman',12,'bold'),width=18,bd=3)
    NameEntry.grid(row=1, column=1,pady=9,padx=10)
    NameEntry.insert(0,f'Glowybar')

    PhoneLabel=Label(DetailsFrame,text='Phone',font=('times new roman',12,'bold'),bg="#06283D",fg='#fff')
    PhoneLabel.grid(row=1,column=2,pady=9,padx=10)
    PhoneEntry=Entry(DetailsFrame,font=('times new roman',12,'bold'),width=18,bd=3)
    PhoneEntry.grid(row=1, column=3,pady=9,padx=10)
    PhoneEntry.insert(0,f'+43 (0)681 20716861')

    billnumberLabel=Label(DetailsFrame,text='Bill_No',font=('times new roman',12,'bold'),bg="#06283D",fg='#fff')
    billnumberLabel.grid(row=1,column=4,pady=9,padx=10)
    billnumberEntry=Entry(DetailsFrame,font=('times new roman',12,'bold'),width=6,bd=3)
    billnumberEntry.grid(row=1, column=5,pady=9,padx=10)

    Searchbutton=Button(DetailsFrame,text='Search',font=('arial',12,'bold'),bg='grey',fg='#fff', command=search_bills)# 
    Searchbutton.grid(row=1,column=6)

    DateLabel=Label(DetailsFrame,text='Date',font=('times new roman',12,'bold'),fg='black')
    DateLabel.grid(row=1,column=7,pady=9,padx=10)
    Date = StringVar()
    current_datetime = datetime.now()
    d1 = current_datetime.strftime("%d-%m-%y %H:%M:%S")
    Date.set(d1)
    DateEntry=Entry(DetailsFrame,textvariable=Date,font=('times new roman',12,'bold'),width=16,bd=3)
    DateEntry.grid(row=1, column=8,pady=9,padx=10)

    ProductsareaLabel=Label(DetailsFrame,text='',font=('times new roman',12,'bold'),bg="#06283D",fg='#06283D')
    ProductsareaLabel.grid(row=2,column=0,padx=True)



    # -- Product Frame -----
    productsFrame=Frame(root0,bg="#06283D")
    productsFrame.pack(fill="both", expand=True)

    # -------- Container ---
    # Container frame for table and scrollbar
    container = Frame(productsFrame,bg="#06283D")
    container.pack(side = LEFT, fill="both", expand=True)

    canvas = Canvas(container, height=10, width=20,bg="#06283D", scrollregion=(0, 0, 20, 10))
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = Scrollbar(container, orient="vertical", command=canvas.yview)
    scrollbar.pack(side="right", fill="y")

    canvas.configure(yscrollcommand=scrollbar.set)

    inner_frame = Frame(canvas,bg="#06283D")

    canvas.create_window((0, 0), window=inner_frame, anchor="nw")

    ProductsareaLabel=Label(inner_frame,text='Produkte',font=('times new roman',13,'bold'),bg="#06283D",fg='#fff',relief=GROOVE)
    ProductsareaLabel.grid(row=0,column=0,padx=True) 
    # Cosmetic Table
    RoseDayLabel=Label(inner_frame, text='Rose Cream Day 1', font=('times new roman',10),bg="#06283D",fg='#fff')
    RoseDayLabel.grid(row=1,column=0,pady=9,padx=10) 
    RoseDayEntry=Entry(inner_frame, font=('times new roman',12,'bold'), width=4, bd=1)
    RoseDayEntry.grid(row=1, column=1)
    RoseDayEntry.insert(0,0)

    RoseNightLabel=Label(inner_frame,text='Rose Cream Night',font=('times new roman',10),bg="#06283D",fg='#fff')
    RoseNightLabel.grid(row=2,column=0,pady=9,padx=10) 
    RoseNightEntry=Entry(inner_frame,font=('times new roman',12,'bold'),width=4,bd=1)
    RoseNightEntry.grid(row=2,column=1)  
    RoseNightEntry.insert(0,0)

    WhiteRoseDayLabel=Label(inner_frame,text='W-Rose Cream Day',font=('times new roman',10),bg="#06283D",fg='#fff')
    WhiteRoseDayLabel.grid(row=3,column=0,pady=9,padx=10)
    WhiteRoseDayEntry=Entry(inner_frame,font=('times new roman',12,'bold'),width=4,bd=1)
    WhiteRoseDayEntry.grid(row=3, column=1)
    WhiteRoseDayEntry.insert(0,0)

    WhiteRoseNightLabel=Label(inner_frame,text='W-Rose Cream Night',font=('times new roman',10),bg="#06283D",fg='#fff')
    WhiteRoseNightLabel.grid(row=4,column=0,pady=9,padx=10)
    WhiteRoseNightEntry=Entry(inner_frame,font=('times new roman',12,'bold'),width=4,bd=1)
    WhiteRoseNightEntry.grid(row=4, column=1)
    WhiteRoseNightEntry.insert(0,0)

    BodylotionLabel=Label(inner_frame,text='Bodylotion',font=('times new roman',10),bg="#06283D",fg='#fff')
    BodylotionLabel.grid(row=5,column=0,pady=9,padx=10)
    BodylotionEntry=Entry(inner_frame,font=('times new roman',12,'bold'),width=4,bd=1)
    BodylotionEntry.grid(row=5, column=1)
    BodylotionEntry.insert(0,0)

    HandCreamLabel=Label(inner_frame,text='HandCream',font=('times new roman',10),bg="#06283D",fg='#fff')
    HandCreamLabel.grid(row=6,column=0,pady=9,padx=10)
    HandCreamEntry=Entry(inner_frame,font=('times new roman',12,'bold'),width=4,bd=1)
    HandCreamEntry.grid(row=6, column=1)
    HandCreamEntry.insert(0,0)

    RoseWaterLabel=Label(inner_frame,text='RoseWater',font=('times new roman',10),bg="#06283D",fg='#fff')
    RoseWaterLabel.grid(row=7,column=0,pady=9,padx=10)
    RoseWaterEntry=Entry(inner_frame,font=('times new roman',12,'bold'),width=4,bd=1)
    RoseWaterEntry.grid(row=7, column=1)
    RoseWaterEntry.insert(0,0)


    # Add more labels and entries here
    canvas.update()
    # Update the scroll region to include the inner frame
    canvas.configure(scrollregion=canvas.bbox("all"))

    # -------- Container 1 ---
    # Container 1 frame for table and scrollbar
    container1 = Frame(productsFrame,bg="#06283D")
    container1.pack(side = LEFT, fill="both", expand=True)

    # Create canvas
    canvas1 = Canvas(container1, height=10, width=20,bg="#06283D", scrollregion=(0, 0, 20, 10))
    canvas1.pack(side="left", fill="both", expand=True)

    # Add a scrollbar
    scrollbar = Scrollbar(container1, orient="vertical", command=canvas1.yview)
    scrollbar.pack(side="right", fill="y")

    # Configure the canvas
    canvas1.configure(yscrollcommand=scrollbar.set)

    # Create another frame inside canvas
    inner_frame1 = Frame(canvas1,bg="#06283D")

    # Add the inner frame to the canvas
    canvas1.create_window((0, 0), window=inner_frame1, anchor="nw")

    ProductsareaLabel=Label(inner_frame1,text='Haare - Wimpel',font=('times new roman',13,'bold'),bg="#06283D",fg='#fff',relief=GROOVE)
    ProductsareaLabel.grid(row=0,column=0,padx=True) 
    # Service Table
    HaarefarbeLabel=Label(inner_frame1, text='Haare-farbe', font=('times new roman',10),bg="#06283D",fg='#fff')
    HaarefarbeLabel.grid(row=1,column=0,pady=9,padx=10) 
    HaarefarbeEntry=Entry(inner_frame1, font=('times new roman',12,'bold'), width=4, bd=1)
    HaarefarbeEntry.grid(row=1, column=1)
    HaarefarbeEntry.insert(0,0)

    DamenhaareLabel=Label(inner_frame1,text='Damen haare cut',font=('times new roman',10),bg="#06283D",fg='#fff')
    DamenhaareLabel.grid(row=2,column=0,pady=9,padx=10) 
    DamenhaareEntry=Entry(inner_frame1,font=('times new roman',12,'bold'),width=4,bd=1)
    DamenhaareEntry.grid(row=2,column=1)  
    DamenhaareEntry.insert(0,0)

    DamenföhlenLabel=Label(inner_frame1,text='D.Waschen+föhlen',font=('times new roman',10),bg="#06283D",fg='#fff')
    DamenföhlenLabel.grid(row=3,column=0,pady=9,padx=10)
    DamenföhlenEntry=Entry(inner_frame1,font=('times new roman',12,'bold'),width=4,bd=1)
    DamenföhlenEntry.grid(row=3, column=1)
    DamenföhlenEntry.insert(0,0)

    Cu_wa_foLabel=Label(inner_frame1,text='C.W.+ Föhlen',font=('times new roman',10),bg="#06283D",fg='#fff')
    Cu_wa_foLabel.grid(row=4,column=0,pady=9,padx=10)
    Cu_wa_foEntry=Entry(inner_frame1,font=('times new roman',12,'bold'),width=4,bd=1)
    Cu_wa_foEntry.grid(row=4, column=1)
    Cu_wa_foEntry.insert(0,0)

    WimpernVLabel=Label(inner_frame1,text='wimpern verlängerung',font=('times new roman',10),bg="#06283D",fg='#fff')
    WimpernVLabel.grid(row=5,column=0,pady=9,padx=10)
    WimpernVEntry=Entry(inner_frame1,font=('times new roman',12,'bold'),width=4,bd=1)
    WimpernVEntry.grid(row=5, column=1)
    WimpernVEntry.insert(0,0)

    HairCutMLabel=Label(inner_frame1,text='Hair Cut Men',font=('times new roman',10),bg="#06283D",fg='#fff')
    HairCutMLabel.grid(row=6,column=0,pady=9,padx=10)
    HairCutMEntry=Entry(inner_frame1,font=('times new roman',12,'bold'),width=4,bd=1)
    HairCutMEntry.grid(row=6, column=1)
    HairCutMEntry.insert(0,0)

    HairWashMLabel=Label(inner_frame1,text='Hair Wash Men',font=('times new roman',10),bg="#06283D",fg='#fff')
    HairWashMLabel.grid(row=7,column=0,pady=9,padx=10)
    HairWashMEntry=Entry(inner_frame1,font=('times new roman',12,'bold'),width=4,bd=1)
    HairWashMEntry.grid(row=7, column=1)
    HairWashMEntry.insert(0,0)


    # Add more labels and entries here
    canvas1.update()
    # Update the scroll region to include the inner frame
    canvas1.configure(scrollregion=canvas1.bbox("all"))

    # -------- Container 2 ---
    # Container 2 frame for table and scrollbar
    container2 = Frame(productsFrame,bg="#06283D")
    container2.pack(side = LEFT, fill="both", expand=True)

    # Create canvas
    canvas2 = Canvas(container2, height=10, width=20,bg="#06283D", scrollregion=(0, 0, 20, 10))
    canvas2.pack(side="left", fill="both", expand=True)

    # Add a scrollbar
    scrollbar = Scrollbar(container2, orient="vertical", command=canvas2.yview)
    scrollbar.pack(side="right", fill="y")

    # Configure the canvas
    canvas2.configure(yscrollcommand=scrollbar.set)

    # Create another frame inside canvas
    inner_frame2 = Frame(canvas2,bg="#06283D")

    # Add the inner frame to the canvas
    canvas2.create_window((0, 0), window=inner_frame2, anchor="nw")

    ProductsareaLabel=Label(inner_frame2,text='Nagel pflege',font=('times new roman',13,'bold'),bg="#06283D",fg='#fff',relief=GROOVE)
    ProductsareaLabel.grid(row=0,column=0,padx=True)
    # Service Table
    AuffüllenLabel=Label(inner_frame2, text='Auffüllen', font=('times new roman',10),bg="#06283D",fg='#fff')
    AuffüllenLabel.grid(row=1,column=0,pady=9,padx=10) 
    AuffüllenEntry=Entry(inner_frame2, font=('times new roman',12,'bold'), width=4, bd=1)
    AuffüllenEntry.grid(row=1, column=1)
    AuffüllenEntry.insert(0,0)

    NeueSetLabel=Label(inner_frame2,text='NeueSet',font=('times new roman',10),bg="#06283D",fg='#fff')
    NeueSetLabel.grid(row=2,column=0,pady=9,padx=10) 
    NeueSetEntry=Entry(inner_frame2,font=('times new roman',12,'bold'),width=4,bd=1)
    NeueSetEntry.grid(row=2,column=1)  
    NeueSetEntry.insert(0,0)

    ShellachLabel=Label(inner_frame2,text='Shellach',font=('times new roman',10),bg="#06283D",fg='#fff')
    ShellachLabel.grid(row=3,column=0,pady=9,padx=10)
    ShellachEntry=Entry(inner_frame2,font=('times new roman',12,'bold'),width=4,bd=1)
    ShellachEntry.grid(row=3, column=1)
    ShellachEntry.insert(0,0)

    BedechungLabel=Label(inner_frame2,text='Bedechung',font=('times new roman',10),bg="#06283D",fg='#fff')
    BedechungLabel.grid(row=4,column=0,pady=9,padx=10)
    BedechungEntry=Entry(inner_frame2,font=('times new roman',12,'bold'),width=4,bd=1)
    BedechungEntry.grid(row=4, column=1)
    BedechungEntry.insert(0,0)

    PedikureLabel=Label(inner_frame2,text='Pedikure',font=('times new roman',10),bg="#06283D",fg='#fff')
    PedikureLabel.grid(row=5,column=0,pady=9,padx=10)
    PedikureEntry=Entry(inner_frame2,font=('times new roman',12,'bold'),width=4,bd=1)
    PedikureEntry.grid(row=5, column=1)
    PedikureEntry.insert(0,0)

    ManikureLabel=Label(inner_frame2,text='Manikure',font=('times new roman',10),bg="#06283D",fg='#fff')
    ManikureLabel.grid(row=6,column=0,pady=9,padx=10)
    ManikureEntry=Entry(inner_frame2,font=('times new roman',12,'bold'),width=4,bd=1)
    ManikureEntry.grid(row=6, column=1)
    ManikureEntry.insert(0,0)

    AcrylEntLabel=Label(inner_frame2,text='Acryl Entfernen',font=('times new roman',10),bg="#06283D",fg='#fff')
    AcrylEntLabel.grid(row=7,column=0,pady=9,padx=10)
    AcrylEntEntry=Entry(inner_frame2,font=('times new roman',12,'bold'),width=4,bd=1)
    AcrylEntEntry.grid(row=7, column=1)
    AcrylEntEntry.insert(0,0)


    # Add more labels and entries here
    canvas1.update()
    # Update the scroll region to include the inner frame
    canvas2.configure(scrollregion=canvas1.bbox("all"))


    # --- Bill Area -----
    billframe=Frame(productsFrame,bd=5,relief=GROOVE)
    billframe.pack(side = TOP, fill="both", expand=True)
    # billframe.grid(row=0,column=1,padx=10)
    billareaLabel=Label(billframe,text='Bill Area',font=('times new roman',12,'bold'),bg="#06283D",fg='#fff',bd=3,relief=GROOVE)
    billareaLabel.pack(side = TOP, fill="both", expand=True)

    scrollbar=Scrollbar(billframe,orient=VERTICAL)
    scrollbar.pack(side=RIGHT,fill=Y)
    textarea=Text(billframe,height=25,width=50,yscrollcommand=scrollbar.set) 
    textarea.pack()
    scrollbar.config(command=textarea.yview)

    # -- Total - Tax - Button ---
    billmenuFrame=LabelFrame(root0,text='Bill Menu',font=('times new roman',13,'bold'),bg="#06283D",fg='#fff',bd=8,relief=GROOVE)
    billmenuFrame.pack(fill=X)

    ProductspriceLabel=Label(billmenuFrame,text='Produkte Price',font=('tomes new roman',12,'bold'),bg='#06283D',fg='white')
    ProductspriceLabel.grid(row=0,column=0,padx=10,pady=9,sticky='w')
    ProductspriceEntry=Entry(billmenuFrame,font=('times new roman',13,'bold'),width=10,bd=3)
    ProductspriceEntry.grid(row=0,column=1,padx=10,pady=9)

    HaarepriceLabel=Label(billmenuFrame,text='Haare Pflege',font=('tomes new roman',12,'bold'),bg='#06283D',fg='white')
    HaarepriceLabel.grid(row=1,column=0,padx=10,pady=9,sticky='w')
    HaarepriceEntry=Entry(billmenuFrame,font=('times new roman',13,'bold'),width=10,bd=3)
    HaarepriceEntry.grid(row=1,column=1,padx=10,pady=9)

    NagelpriceLabel=Label(billmenuFrame,text='Nagel Pflege',font=('tomes new roman',12,'bold'),bg='#06283D',fg='white')
    NagelpriceLabel.grid(row=2,column=0,padx=10,pady=9,sticky='w')
    NagelpriceEntry=Entry(billmenuFrame,font=('times new roman',13,'bold'),width=10,bd=3)
    NagelpriceEntry.grid(row=2,column=1,padx=10,pady=9)

    ProductstaxLabel=Label(billmenuFrame,text='Produkte tax',font=('tomes new roman',12,'bold'),bg='#06283D',fg='white')
    ProductstaxLabel.grid(row=0,column=3,padx=10,pady=9,sticky='w')
    ProductstaxEntry=Entry(billmenuFrame,font=('times new roman',13,'bold'),width=9,bd=3)
    ProductstaxEntry.grid(row=0,column=4,padx=10,pady=9)

    HaaretaxLabel=Label(billmenuFrame,text='Haare pflege tax',font=('tomes new roman',12,'bold'),bg='#06283D',fg='white')
    HaaretaxLabel.grid(row=1,column=3,padx=10,pady=9,sticky='w')
    HaaretaxEntry=Entry(billmenuFrame,font=('times new roman',13,'bold'),width=9,bd=3)
    HaaretaxEntry.grid(row=1,column=4,padx=10,pady=9)

    NageltaxLabel=Label(billmenuFrame,text='Nagel pflege tax',font=('tomes new roman',12,'bold'),bg='#06283D',fg='white')
    NageltaxLabel.grid(row=2,column=3,padx=10,pady=9,sticky='w')
    NageltaxEntry=Entry(billmenuFrame,font=('times new roman',13,'bold'),width=9,bd=3)
    NageltaxEntry.grid(row=2,column=4,padx=10,pady=9)

    buttonFrame=Frame(billmenuFrame,bd=8,bg='#06283D',relief=GROOVE)
    buttonFrame.grid(row=0,column=5,rowspan=6,padx=True)

    Check_outbutton=Button(buttonFrame,text='Check-out',font=('arial',15,'bold'),bg='grey',fg='white',bd=3,width=8,command=Check_outbutton) # 
    Check_outbutton.grid(row=0,column=0,padx=10,pady=10)

    billbutton=Button(buttonFrame,text='bill',font=('arial',15,'bold'),bg='grey',fg='white',bd=3,width=5,command=bill_Area)# 
    billbutton.grid(row=0,column=1,padx=10,pady=10)

    Emailbutton=Button(buttonFrame,text='Email',font=('arial',15,'bold'),bg='grey',fg='white',bd=3,width=5, command=send_email)# 
    Emailbutton.grid(row=0,column=2,padx=10,pady=10)

    Printbutton=Button(buttonFrame,text='Print',font=('arial',15,'bold'),bg='grey',fg='white',bd=3,width=5, command=print_bill)# 
    Printbutton.grid(row=0,column=3,padx=10,pady=10)

    Clearbutton=Button(buttonFrame,text='Clear',font=('arial',15,'bold'),bg='grey',fg='white',bd=3,width=5,command=clear)# 
    Clearbutton.grid(row=0,column=4,padx=10,pady=10)

    Settingbutton=Button(buttonFrame,text='Setting',font=('arial',15,'bold'),bg='grey',fg='white',bd=3,width=6)
    Settingbutton.grid(row=0,column=5,padx=10,pady=10)

    Exitbutton=Button(buttonFrame,text='Exit',font=('arial',15,'bold'),bg='grey',fg='white',bd=3,width=5,command=Exitbutton)# 
    Exitbutton.grid(row=0,column=6,padx=10,pady=10)
    
    root0.mainloop()

# ############################# end root0  #########################################
# --- View dữ liệu trong Excel Sk_Data ---
def view_data():
    # Tạo cửa sổ mới
    view_window = Tk.Toplevel(root)
    view_window.title("Xem dữ liệu")
    view_window.geometry("700x500")

    # Đọc dữ liệu từ file Excel
    file1 = "SK_Daten.xlsx"
    workbook = openpyxl.load_workbook(file1)
    sheet1 = workbook.active

    # Hiển thị dữ liệu trên cửa sổ mới
    for row in sheet1.iter_rows(values_only=True):
        row_data = " | ".join(str(cell) for cell in row)
        label1 = Label(view_window, text=row_data, anchor="w")
        label1.pack(anchor="w", padx=10, pady=5)

    # Đóng file Excel
    workbook.close()
    
    view_window.option_add("*justify", "left")
# -------- Hết -----

def view_Point():
    # Tạo cửa sổ mới
    view_window = Tk.Toplevel(root)
    view_window.title("Xem dữ liệu")
    view_window.geometry("400x500")

    # Đọc dữ liệu từ file Excel
    file2 = "SK_Point.xlsx"
    workbook = openpyxl.load_workbook(file2)
    sheet2 = workbook.active

    # Hiển thị dữ liệu trên cửa sổ mới
    for row in sheet2.iter_rows(values_only=True):
        row_data = " | ".join(str(cell) for cell in row)
        label2 = Label(view_window, text=row_data, anchor="w")
        label2.pack(anchor="w", padx=10, pady=5)

    # Đóng file Excel
    workbook.close()
    
    view_window.option_add("*justify", "left")
    
# ---- Hết --------

# Lấy đường dẫn tuyệt đối của thư mục hiện tại
current_dir = os.path.dirname(os.path.abspath(__file__))

# Xác định đường dẫn đến tệp ico
icon_path = os.path.join(current_dir, "Images", "app.ico")

# Sử dụng biểu tượng


root = Tk.Tk()
root.title("Registieren System")
root.iconbitmap(icon_path)
# root.iconbitmap('Images/app.ico')
root.geometry("1250x730+210+100")

background = "#06283D"  
framebg = "#EDEDED"
framefg = "#06283D"
root.config(bg=background)

file1 = pathlib.Path('SK_Daten.xlsx')
if not file1.exists():
    file1 = Workbook()
    sheet1 = file1.active
    sheet1['A1'] = "Kundenbarcode"
    sheet1['B1'] = "Namen"
    sheet1['C1'] = "Geburtstag"
    sheet1['D1'] = "Geschlecht"
    sheet1['E1'] = "Punkte"
    sheet1['F1'] = "Phone"
    sheet1['G1'] = "Email"
    sheet1['H1'] = "Adresse"
    sheet1['I1'] = "Reg_Date"
    file1.save('SK_Daten.xlsx')

file2 = pathlib.Path('SK_Point.xlsx')
if not file2.exists():
    file2 = Workbook()
    sheet2 = file2.active
    sheet2['A1'] = "Reg_No"
    sheet2['B1'] = "Produkte"
    sheet2['C1'] = "Datum"
    sheet2['D1'] = "Wert"
    sheet2['E1'] = "Kd_Code"
    file2.save('SK_Point.xlsx')

######################################################    
def Exit():
	root.destroy()
######################################################	
def showimage():
    global filename 
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
			                            title="Select image file",filetype=(("JPG File","*.jpg"),
					                                                        ("PNG File","*.png"),
					                                                        ("All files","*.txt")))
     
    img = (Image.open(filename))
    resized_image= img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2
    
#####################################################################    
def Registration_no(): 
	file2=openpyxl.load_workbook('SK_Point.xlsx')
	sheet2=file2.active
	row=sheet2.max_row

	max_row_value=sheet2.cell(row=row,column=1).value
	
	try:
		Registration.set(max_row_value+1)
	except:
		Registration.set("1")
#######################################################################
	
def Save_Gutschein_Bonus(): # xuống dưới code của Button Save thêm command=Save
    R1 = Registration.get()
    S1 = Produkte.get()
    T1 = Wert.get()
    U1 = Kd_Code.get()
    V1 = Date1.get()
    
    if U1 == "" or S1 == "" or T1 == "" or V1 == "":
        messagebox.showerror("Error","Thiếu thông tin !")
    else:
        file2 = openpyxl.load_workbook('SK_Point.xlsx')
        sheet2 = file2.active
        sheet2.cell(column=1, row=sheet2.max_row +1, value=R1)
        sheet2.cell(column=2, row=sheet2.max_row, value=S1)
        sheet2.cell(column=3, row=sheet2.max_row, value=V1)
        sheet2.cell(column=4, row=sheet2.max_row, value=T1)
        sheet2.cell(column=5, row=sheet2.max_row, value=U1)
        
        file2.save("SK_Point.xlsx")
        
        messagebox.showinfo("Info","Lưu thành công")
        
        Clear()
        Registration_no()


########################

def Save_Kunden_Data(): # xuống dưới code của Button Save thêm command=Save
    R1 = Kundenbarcode.get()
    N1 = Namen.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("error", "Select Gender !")

    D2 = DOB.get()
    D1 = Date.get()
    P1 = Phone.get()
    E1 = Email.get()
    J1 = label_punkte.cget("text") # Chat Gpt
    M1 = Adresse.get()             # adresse_entry.get("1.0", "end-1c") # Chat Gpt

    if R1 == "" or N1 == "" or D2 == "" or P1 == "" or E1 == "" or M1 == "" or J1 == "":
        messagebox.showerror("error", "Thieu thong tin !")
    else:
        file1 = openpyxl.load_workbook('SK_Daten.xlsx')
        sheet1 = file1.active
        
        for row in range(1, sheet1.max_row + 1):
            if sheet1.cell(column=1, row=row).value == R1:
                messagebox.showerror("error", "Mã khách hàng đã tồn tại!")
                return
            
        sheet1.cell(column=1, row=sheet1.max_row + 1, value=R1)
        sheet1.cell(column=2, row=sheet1.max_row, value=N1)
        sheet1.cell(column=3, row=sheet1.max_row, value=D2)
        sheet1.cell(column=4, row=sheet1.max_row, value=G1)
        sheet1.cell(column=5, row=sheet1.max_row, value=J1)
        sheet1.cell(column=6, row=sheet1.max_row, value=P1)
        sheet1.cell(column=7, row=sheet1.max_row, value=E1)
        sheet1.cell(column=8, row=sheet1.max_row, value=M1)
        sheet1.cell(column=9, row=sheet1.max_row, value=D1)

        file1.save(r'SK_Daten.xlsx')

        try:
            img.save("Kd_Images/" + str(R1) + ".jpg")
        except:
            messagebox.showerror("Error", "chua co hinh anh !")

        messagebox.showinfo("info", "luu thanh cong !")
        
        
        label_punkte.config(text="0") # Chat Gpt

        Clear()
        
#################### Search #################
# tạo hiệu ứng cho mục Search ----
def search():
    text = Search.get()

    if text == "":
        messagebox.showerror("Lỗi", "Chưa nhập thông tin tìm kiếm!")
        return

    # Xóa thông tin trước đó trên giao diện
        Clear()
    
        # Bật Button Update
        Update_button.config(state='normal')


    # Ẩn button Save_Kunden_DataButton
    Save_Kunden_DataButton.config(state='disable')

    file1 = openpyxl.load_workbook("SK_Daten.xlsx")
    sheet1 = file1.active

    found = False  # Biến đánh dấu xem có tìm thấy thông tin hay không

    for row in sheet1.rows:
        # Kiểm tra nếu giá trị trong cột đầu tiên của hàng chứa thông tin bạn đang tìm kiếm
        if text in str(row[0].value):
            # Lưu thông tin hàng và đánh dấu là đã tìm thấy
            name = row[0]
            Kundenbarcode_number = row[0].row
            found = True
            break

    if found:
        try:
            print(str(name))
        except:
            messagebox.showerror("Lỗi", "Số khách hàng không hợp lệ!")

        x1 = sheet1.cell(row=int(Kundenbarcode_number), column=1).value
        x2 = sheet1.cell(row=int(Kundenbarcode_number), column=2).value
        x3 = sheet1.cell(row=int(Kundenbarcode_number), column=3).value
        x4 = sheet1.cell(row=int(Kundenbarcode_number), column=4).value
        x5 = sheet1.cell(row=int(Kundenbarcode_number), column=5).value
        x6 = sheet1.cell(row=int(Kundenbarcode_number), column=6).value
        x7 = sheet1.cell(row=int(Kundenbarcode_number), column=7).value
        x8 = sheet1.cell(row=int(Kundenbarcode_number), column=8).value
        x9 = sheet1.cell(row=int(Kundenbarcode_number), column=9).value

        # Đặt giá trị cho các biến StringVar
        Kundenbarcode.set(x1)
        Namen.set(x2)
        DOB.set(x3)
        if x4 == 'Frau':
            R2.select()
        else:
            R1.select()
        Punkte.set(x5)
        Phone.set(x6)
        Email.set(x7)
        Adresse.set(x8)
        Date.set(x9)

        # Load ảnh
        try:
            img = Image.open("Kd_Images/" + str(x1) + ".jpg")
            resized_image = img.resize((190, 190))
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image = photo2
        except FileNotFoundError:
            messagebox.showerror("Lỗi", "Không tìm thấy ảnh!")

    else:
        # Hiển thị thông báo lỗi nếu không tìm thấy
        messagebox.showerror("Lỗi", "Không tìm thấy thông tin!")

    # Đóng file Excel
    file1.close()

#######################  Update ################
def Update():
    
    R1 = Kundenbarcode.get()
    N1 = Namen.get()
    selection()
    G1 = gender
    D2 = DOB.get()
    D1 = Date.get()
    P1 = Phone.get()
    E1 = Email.get()
    J1 = label_punkte.cget("text")
    M1 = Adresse.get()

    file1 = openpyxl.load_workbook("SK_Daten.xlsx")
    sheet1 = file1.active

    for row in sheet1.rows:
        if row[0].value == R1:
            name = row[0]
            print(str(name))
            Kundenbarcode_position = str(name)[14:-1]
            Kundenbarcode_number = str(name)[15:-1]

            print(Kundenbarcode_number)

    sheet1.cell(column=1, row=int(Kundenbarcode_number), value=R1)
    sheet1.cell(column=2, row=int(Kundenbarcode_number), value=N1)
    sheet1.cell(column=3, row=int(Kundenbarcode_number), value=D2)
    sheet1.cell(column=4, row=int(Kundenbarcode_number), value=G1)
    sheet1.cell(column=5, row=int(Kundenbarcode_number), value=J1)
    sheet1.cell(column=6, row=int(Kundenbarcode_number), value=P1)
    sheet1.cell(column=7, row=int(Kundenbarcode_number), value=E1)
    sheet1.cell(column=8, row=int(Kundenbarcode_number), value=M1)
    sheet1.cell(column=9, row=int(Kundenbarcode_number), value=D1)

    file1.save(r'SK_Daten.xlsx')
    try:
        img.save("Kd_Images/" + str(R1) + ".jpg")
    except:
        pass
    messagebox.showinfo("Update", "Update Successfully !")
    
    label_punkte.config(text="0")
    
    Clear()
#####################################################################
def Clear():
    global img
    Namen.set('')
    Kd_Code.set('')
    DOB.set('')
    Kundenbarcode.set('')
    Phone.set('')
    Email.set('')
    Adresse.set('')
    Punkte.set('')
    # Datum.set('')
    Kd_Code.set('')
    Wert.set('')
    Search.set('')
    Produkte.set("Select Produkte")
        
    Registration_no()
        
    label_punkte.config(text="0")
        
    # adresse_entry.delete("1.0", "end")
        
    Save_Kunden_DataButton.config(state = 'normal')
        
    img1=PhotoImage(file='Images/Bild.png')
    lbl.config(image=img1)
    lbl.image=img1
        
    img=""
        
    # Update_button.config(state='disable')
        
    label_punkte.config(foreground="black")

#####################################################################

# ------- Point ----------
def Point():
    global CurrentPoint
    NewPoint = int(Punkte.get())
    # Lấy giá trị hiện tại của label_punkte và chuyển đổi sang kiểu số nguyên
    CurrentPoint = int(label_punkte.cget("text"))
    TotalPoint = NewPoint + CurrentPoint         

    # Kiểm tra điều kiện để thay đổi màu
    if TotalPoint > 99:
        label_punkte.config(foreground="red")
        # Thực hiện thay đổi màu sang đỏ
        # Ví dụ: label_punkte.config(foreground="red")
        pass
    else:
        # Giữ nguyên màu
        label_punkte.config(foreground="black")
        pass
    
    # Cập nhật giá trị mới lên label_punkte
    label_punkte.config(text=str(TotalPoint))

       
####################################################################

#-------- Gender -----------
def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Mann"
    else:
        gender = "Frau"
# ------ gọi giao diện preis-Rechnen ----       
# def open_recapp():
    # subprocess.call(["python", "RecApp.py"]) # cần import subprocess !
    

# ----- Frame 1 -----
		
Label(root,text="Email : office@lerose.at - Tel.: +43 699 1927 9271 ", width=10, height=2,bg="Silver", anchor='e').pack(side=TOP,fill=X) 
Label(root,text="NAGELSTUDIO LE", width=30,height=2,bg="#06283D",fg='#fff',font='arial 30 bold').place(x=0,y=40)
Label(root,text="Stammkunden Management", width=30,height=1,bg="#06283D",fg='#fff',font="System 14").place(x=200,y=110)

Search=StringVar()
Entry(root, textvariable=Search, width=10,  bd=5, font="arial 13").place(x=1000,y=80) 
# imageicon3=PhotoImage(file="Images/Search1.png")
Srch=Button(root,text="Search", width=8, bg='Silver', font="arial 10 bold", bd=5, command=search)  
Srch.place(x=1120,y=80) 

Update_button=Button(root, text="Update", width=8, bg='Silver', font="arial 10 bold", bd=5, command=Update).place(x=850,y=80)# , state='disable'
# --- Frame  ------- 
Label(root,text="Kd_code",font="arial 13",fg=framebg,bg=background).place(x=50,y=170) 
Label(root,text="Datum_Reg.",font="arial 13",fg=framebg,bg=background).place(x=300,y=170)

Kundenbarcode=StringVar()
Kcode_entry = Entry(root,textvariable=Kundenbarcode,width=10,font="arial 12")
Kcode_entry.place(x=180,y=170)

Date = StringVar()
today = date.today()
d1 = today.strftime("%d/%m/%y")
#print(d1) # xem ngay thang co dung khong ?
date_entry = Entry(root,textvariable=Date,width=15,font="arial 12")
date_entry.place(x=400,y=170)

Date.set(d1)
# Tạo button để mở cửa sổ xem dữ liệu
#------ Button View --
view_button = Button(root, text="View Kunden Daten ", border=0, bg="#06283D", fg='#fff', command=view_data).place(x=680,y=170)
view_button2 = Button(root, text="- View Bonus - Gutschein", border=0, bg="#06283D", fg='#fff', command=view_Point).place(x=790,y=170)
# view_button.pack() 
# ------  obj 1 ----------
obj=LabelFrame(root,text="Kunden details",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Namen :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=40)
Label(obj,text="Geburtag : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=90)
Label(obj,text="Geschlecht :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=140)
Label(obj,text="Punkte :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=190)

Label(obj,text="Phone :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=40)
Label(obj,text="Email :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=90)
Label(obj,text="Adresse :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=140)
Label(obj,text="Hãy nhập số điểm mới vào ô trống rồi nhấn dấu + để lấy tổng số điểm tích lũy !",font="arial 10",bg=framebg,fg=framefg).place(x=340,y=190)

Namen=StringVar() 
namen_entry = Entry(obj,textvariable=Namen,width=20,font="arial 10")
namen_entry.place(x=160,y=40)

DOB=StringVar() 
dob_entry = Entry(obj,textvariable=DOB,width=20,font="arial 10")
dob_entry.place(x=160,y=90)
radio= IntVar()
R1 = Radiobutton(obj,text="Mann", variable=radio, value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=170,y=140)
R2 = Radiobutton(obj,text="Frau", variable=radio, value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=240,y=140)

Punkte = StringVar()
punkte_entry = Entry(obj, textvariable=Punkte, width=5, font="arial 12 bold")
punkte_entry.place(x=160, y=190)

# Tạo label và đặt giá trị ban đầu là 10
label_punkte = Label(obj, text="0", font="arial 20 bold")
label_punkte.place(x=260, y=185)
# change_color()
# Tạo nút để thực hiện việc cộng giá trị
nut_cong = Button(obj, text=" + ",bg="Silver", command=Point) # , command=Point
nut_cong.place(x=220, y=190)

Phone=StringVar() 
phone_entry = Entry(obj,textvariable=Phone,width=24,font="arial 10")
phone_entry.place(x=630,y=40)

Email=StringVar() 
email_entry = Entry(obj,textvariable=Email,width=24,font="arial 10")
email_entry.place(x=630,y=90)

Adresse = StringVar()
adresse_entry = Entry(obj,textvariable=Adresse, width=24, font="arial 10")
adresse_entry.place(x=630, y=140)

#------- obj 2 -----------
obj2=LabelFrame(root,text="Gutschein - Bonus",font=20,bd=2,width=900,bg=framebg,fg=framefg,height=220,relief=GROOVE)
obj2.place(x=30,y=470) 

Label(obj2,text="Reg_Nr :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=30)
Label(obj2,text="Produkte :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=80)
Label(obj2,text="Im Wert :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=130)

Registration=IntVar()
reg_entry = Entry(obj2,textvariable=Registration,width=20,font="arial 10")
reg_entry.place(x=160,y=30)

Registration_no()

Produkte= Combobox(obj2,value=['Gutschein','Bonus','Rabat 10%','Rabat 20%'],font="Roboto 10",width=17, state="r")
Produkte.place(x=160,y=80) 
Produkte.set("Chọn Sản Phẩm")

Wert=StringVar()
Wert_entry = Entry(obj2,textvariable=Wert,width=20,font="arial 10")
Wert_entry.place(x=160,y=130)

Label(obj2,text="Kd_Code",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=30)
Label(obj2,text="Datum :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=80)
# Label(obj2,text="Datum_P :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=130)

Kd_Code=StringVar()
Kd_Code_entry = Entry(obj2,textvariable=Kd_Code,width=20,font="arial 10")
Kd_Code_entry.place(x=630,y=30)

Date1 = StringVar()
today = date.today()
d3 = today.strftime("%d/%m/%y")
#print(d1) # xem ngay thang co dung khong ?
date1_entry = Entry(obj2,textvariable=Date1,width=16,font="arial 12")
date1_entry.place(x=630,y=80)

Date1.set(d3)

f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="Images/Bild.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0) # chay thu 15

	# Btton, Save, reset, exit
Button(root, text="Upload", width=20, height=2,font="arial 12 bold",bg="Silver", bd=6, command=showimage).place(x=1000, y=380)  
Save_Kunden_DataButton=Button(root, text="Save Daten", width=20, height=2,font="arial 12 bold",bg="Silver",bd=6, command=Save_Kunden_Data)
Save_Kunden_DataButton.place(x=1000, y=450) #, command=Save_Kunden_Data
Button(obj2, text="Save G und B", width=17, height=1,font="arial 10 bold",bg="Silver",bd=6, command=Save_Gutschein_Bonus).place(x=630,y=130) 
Button(root, text="Reset", width=20, height=2,font="arial 12 bold",bg="Silver",bd=6, command=Clear).place(x=1000, y=520) 
Button(root, text="Exit", width=20, height=2,font="arial 12 bold",bg="Silver",bd=3, command=Exit).place(x=1000, y=590)
Button(root, text="Preis-Rechnen", width=20, height=1,font="arial 12 bold",bg="Silver",bd=6, activeforeground="Silver", command=open_recapp).place(x=1000, y=660) 



root.mainloop()


